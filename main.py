#SPICE DISPENSOR GUI
#Created by: Rohan Patel
#Date modified: 05/23/2023

from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import font as tkFont
import fractions
from quantity import fracmes
import openpyxl
import time


#global variables
selectedSpice = ""
container = ""
frac = fractions.Fraction(1,4)
measurement = "tsp"
recipe = {}
savedRecipes = {}
lastMenu = ""
positions = [""] * 20
selectOrLogged = ""
mainOrLogged = ""
selectedRecipe = ""

#global functions
def selected(rName):
    global selectedRecipe
    selectedRecipe = rName
    return selectedRecipe

def changeText(name):
    global selectedSpice
    selectedSpice = name
    return selectedSpice

def setRecipe(rName):
    global savedRecipes
    global recipe
    recipe = savedRecipes[rName]

def addFrac():
        global frac
        frac = frac + fractions.Fraction(1,4)
        return frac

def subFrac():
        global frac
        if frac == fractions.Fraction(1,4):
            return frac
        else:
            frac = frac - fractions.Fraction(1,4)
        return frac

def saveSelect():
    global recipe
    global measurement
    num = frac.numerator/frac.denominator
    amt = measurement
    quant = fracmes(num, amt)
    recipe[selectedSpice] = quant

def removeSpice(k):
    global recipe
    del recipe[k]

def sorl(menu):
    global selectOrLogged
    selectOrLogged = menu

def morl(menu):
    global mainOrLogged
    mainOrLogged = menu

def last(menu):
    global lastMenu
    lastMenu = menu

def changeToTsp():
    global measurement
    measurement = "tsp"
    return measurement

def changeToTbsp():
    global measurement
    measurement = "tbsp"
    return measurement

def reduceFrac(frac):
    wn =  frac.numerator // frac.denominator
    r = frac.numerator % frac.denominator
    if r == 0:
        return str(wn)
    elif wn == 0:
        return str(r) + "/" + str(frac.denominator)
    else:
        return str(wn) + " " + str(r) + "/" + str(frac.denominator)
    
def saveRecipe():
    global recipe
    file = "Recipes.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheets = workbook.sheetnames
    sheets.pop(0)
    index = len(sheets) + 1
    workbook.create_sheet('Recipe ' + str(index))
    workbook.active = workbook['Recipe ' + str(index)]
    i = 1
    for k,v in recipe.items():
        workbook.active["A" + str(i)] = k
        workbook.active["B" + str(i)] = v.dec
        workbook.active["C" + str(i)] = v.amt
        i += 1
    workbook.save("Recipes.xlsx")

def editRecipe():
    global recipe
    global selectedRecipe
    file = "Recipes.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheets = workbook.sheetnames
    sheets.pop(0)
    workbook.active = workbook[selectedRecipe]
    workbook.active.delete_rows(len(recipe))
    i = 1
    for k,v in recipe.items():
        workbook.active["A" + str(i)] = k
        workbook.active["B" + str(i)] = v.dec
        workbook.active["C" + str(i)] = v.amt
        i += 1
    workbook.save("Recipes.xlsx")

def changePositions(drop1, drop2, drop3, drop4, drop5, temp):
    global positions
    toChange = [drop1.get(), drop2.get(), drop3.get(), drop4.get(), drop5.get()]

    k = TRUE
    for i in range(len(toChange)):
        if toChange.count(toChange[i]) > 1:
            k = FALSE
            break

    if k == FALSE:
        temp.config(text = "Spices cannot be the same, please change your selection")
        temp['font'] = tkFont.Font(family='Helvetica', size=20, weight='bold')
    else:
        temp.config(text = "Your changes have been saved!")
        temp['font'] = tkFont.Font(family='Helvetica', size=30, weight='bold')
        j = 0
        for spice in toChange:
            index = positions.index(spice)
            positions[j],positions[index] = positions[index],positions[j]
            j += 1
        savePositions()

def savePositions():
    global positions
    file = "car_loc.xlsx"
    workbook = openpyxl.load_workbook(file)
    workbook.active = workbook['Spice Locations']
    for i in range(1,21):
        workbook.active.cell(row=i, column=1).value = positions[i-1]
    workbook.save("car_loc.xlsx")

def deleteRecipe(num):
    file = "Recipes.xlsx"
    workbook = openpyxl.load_workbook(file)
    std = workbook['Recipe ' + str(num)]
    workbook.remove(std)
    workbook.save("Recipes.xlsx")

    sheets = workbook.sheetnames
    sheets.pop(0)

    for i in range(len(sheets)):
        name = workbook[sheets[i]]
        name.title = 'Recipe ' + str(i + 1)
    
    workbook.save("Recipes.xlsx")

def loadPositions():
    global positions
    file = "car_loc.xlsx"
    workbook = openpyxl.load_workbook(file)
    sh = workbook.active
    for i in range(1,21):
        spice = sh.cell(row=i, column=1)
        positions[i-1] = spice.value

def runDispense(label, button, cont, self):
    global recipe
    label.config(text = "Dispensing...")
    button.configure(state='disabled', disabledforeground='white', bg='maroon')
    self.update()
    time.sleep(1)
    #main_func.dispense(recipe)
    cont.show_frame(mainMenu)

#Start GUI
class main(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        global container
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        
        #default settings for GUI
        self.geometry("1024x600")
        self.title("Team Spice GUI")
        #self.attributes('-fullscreen',True)
        self.maxsize(1024,600)
        self.minsize(1024,600)
        #removes mouse cursor for touchscreen
        #self.config(cursor="none")
        

        self.frames = {}
        self.frame = None #new

        for F in (mainMenu, selectMenu, positionMenu, loggedMenu, spiceDispense, editMenu, placeContainer):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(mainMenu)

    def show_frame(self, cont):
        global container

        #update spiceDispense frame with new data
        self.frames[cont] = cont(parent=container, controller=self)
        self.frames[cont].grid(row=0, column=0, sticky="nsew")
        frame = self.frames[cont]
        frame.tkraise()

class mainMenu(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)

        font = tkFont.Font(family='Helvetica', size=36, weight='bold')
        mainMenu.configure(self, background='white')
        
        #main menu screen setup
        Select = Button(self, text = "Select", height=1, width=15, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [recipe.clear(), morl(mainMenu), controller.show_frame(selectMenu)])
        Select['font'] = font
        Select.place(x=512, y=150, anchor=CENTER)

        Position = Button(self, text = "Position", height=1, width=15, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: controller.show_frame(positionMenu))
        Position['font'] = font
        Position.place(x=512, y=300, anchor=CENTER)
        
        LR = Button(self, text = "Logged Recipes", height=1, width=15, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: controller.show_frame(loggedMenu))
        LR['font'] = font
        LR.place(x=512, y=450, anchor=CENTER)

class selectMenu(tk.Frame):
    def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)
            font1= tkFont.Font(family='Helvetica', size=15, weight ='bold')
            font2 = tkFont.Font(family='Helvetica', size=23, weight='bold')
            font3 = tkFont.Font(family='Helvetica', size=22, weight='bold')
            font4 = tkFont.Font(family='Helvetica', size=30, weight='bold')
            selectMenu.configure(self, background='white')

            loadPositions()

            instruct = Label(self, text = "Select a spice", background='white')
            instruct['font'] = font4
            instruct.place(x=512, y=40, anchor=CENTER)

            goBack = Button(self, text = "⬅", height=1, width=7, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [controller.show_frame(mainOrLogged)])
            goBack['font'] = font1
            goBack.place(x=20, y=10)

            global positions
            buttons = []
            x = 512
            y = 120
            for i in range(5):
                name = positions[i]
                spice1 = Button(self, text = name, height=1, width=11, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda name=name: [changeText(name), last(selectMenu), controller.show_frame(spiceDispense)])
                spice1['font'] = font2
                spice1.place(x=x, y=y, anchor=CENTER)
                buttons.append(spice1)
                if (i+1) % 5 == 0:
                    y = 120
                    x += 250
                else:
                    y += 85

            for i in range(len(buttons)):
                if buttons[i].cget('text') in recipe:
                    buttons[i].configure(state='disabled', disabledforeground='white', bg='maroon')
                

            if mainOrLogged == mainMenu:
                edit = Button(self, text = "📝 Edit", height=1, width=8, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [sorl(selectMenu), last(selectMenu), controller.show_frame(editMenu)])
                edit['font'] = font3
                edit.place(x=550, y=520)

                #add motor spin to this button
                dispense = Button(self, text = "Dispense", height=1, width=8, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [controller.show_frame(placeContainer)])
                # import main_func
                # command=lambda: [main_func.dispense(recipe)]
                dispense['font'] = font3
                dispense.place(x=320, y=520)
            

class spiceDispense(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        font1= tkFont.Font(family='Helvetica', size=15, weight ='bold')
        font2 = tkFont.Font(family='Helvetica', size=36, weight='bold')
        font3= tkFont.Font(family='Helvetica', size=22, weight ='bold')
        font4= tkFont.Font(family='Helvetica', size=30, weight ='bold')

        spiceDispense.configure(self, background='white')

        goBack = Button(self, text = "⬅", height=1, width=7, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: controller.show_frame(lastMenu))
        goBack['font'] = font1
        goBack.place(x=20, y=10)


        if mainOrLogged == editMenu:
            check = Button(self, text = "☑", height=1, width=7, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [saveSelect(), controller.show_frame(editMenu)])
            check['font'] = font1
            check.place(x=910, y=10)
        elif mainOrLogged == mainMenu:
            check = Button(self, text = "☑", height=1, width=7, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [saveSelect(), controller.show_frame(lastMenu)])
            check['font'] = font1
            check.place(x=910, y=10)

        spiceName = Label(self, text = selectedSpice, background='white')
        spiceName['font'] = font2
        spiceName.place(x=512, y=150, anchor=CENTER)

        global frac
        global measurement
        frac = fractions.Fraction(1,4)
        measurement = "tsp"
        amt = Label(self, text = str(frac) + " " + measurement + ".", width = 8, borderwidth = 1, relief="solid", background='white')
        amt['font'] = font2
        amt.place(x=512, y=250, anchor=CENTER)

        plus = Button(self, text = "+", height=1, width=3, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [addFrac() ,amt.configure(text = reduceFrac(frac) + " " + measurement + ".")])
        plus['font'] = font3
        plus.place(x=675, y=250, anchor=CENTER)

        minus = Button(self, text = "-", height=1, width=3, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [subFrac(), amt.configure(text = reduceFrac(frac) + " " + measurement + ".")])
        minus['font'] = font3
        minus.place(x=348, y=250, anchor=CENTER)

        tsp = Button(self, text = "tsp.", height=1, width=5, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: amt.configure(text = reduceFrac(frac) + " " + changeToTsp() + "."))
        tsp['font'] = font4
        tsp.place(x=388, y=350, anchor=CENTER)

        tbsp = Button(self, text = "tbsp.", height=1, width=5, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: amt.configure(text = reduceFrac(frac) + " " + changeToTbsp() + "."))
        tbsp['font'] = font4
        tbsp.place(x=635, y=350, anchor=CENTER)

class editMenu(tk.Frame):
    def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)

            font1= tkFont.Font(family='Helvetica', size=15, weight ='bold')
            font2 = tkFont.Font(family='Helvetica', size=22, weight='bold')
            editMenu.configure(self, background='white')

            goBack = Button(self, text = "⬅", height=1, width=7, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [controller.show_frame(selectOrLogged)])
            goBack['font'] = font1
            goBack.place(x=20, y=10)

            table = Label(self, relief="solid", background='white')
            table.place(x=75, y=75)
            
            #Create table
            r = 1
            c = 0

            #Column Headers
            table.e = Entry(table, width = 15, font=('Arial', 35, 'bold'))
            table.e.grid(row = 0,column = 0)
            table.e.insert(END, "SPICE NAME")
            table.e.configure(state = DISABLED, disabledbackground = 'white', disabledforeground= 'black')

            table.e = Entry(table, width = 15, font=('Arial', 35, 'bold'))
            table.e.grid(row = 0,column = 1)
            table.e.insert(END, "QUANTITY")
            table.e.configure(state = DISABLED, disabledbackground = 'white', disabledforeground= 'black')
            
            #for button
            x = 900
            y = 165
            
            #Fill table with data from recipe dictionary
            buttons = []
            for k,v in recipe.items():
                table.e = Entry(table, width = 15, font=('Arial', 35))
                table.e.grid(row = r,column = c)
                table.e.insert(END, k)
                table.e.configure(state = DISABLED, disabledbackground = 'white', disabledforeground= 'black')

                Edit = Button(self, text = "📝", height=1, width=3, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda k=k: [changeText(k), last(editMenu), controller.show_frame(spiceDispense)])
                Edit['font'] = font1
                Edit.place(x=x, y=y, anchor = CENTER)
                buttons.append(Edit)
                x += 60

                Delete = Button(self, text = "     🗑️", height=1, width=3, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda k=k: [removeSpice(k), controller.show_frame(editMenu)])
                Delete['font'] = font1
                Delete.place(x=x, y=y, anchor = CENTER)
                buttons.append(Delete)
                x -= 60
                y += 57

                c += 1
                vfrac = fractions.Fraction(v.dec)
                table.e = Entry(table, width = 15, font=('Arial', 35))
                table.e.grid(row = r, column = c)
                table.e.insert(END, reduceFrac(vfrac) + " " + v.amt + ".")
                table.e.configure(state = DISABLED, disabledbackground = 'white', disabledforeground= 'black')
                c = 0
                r += 1
            
            if selectOrLogged == selectMenu and len(recipe) > 1:
                temp=Label(self , text="Your recipe has been logged!", background='white')
                temp['font'] = font2

                save = Button(self, text = "Save Recipe", height=1, width=10, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [saveRecipe(), temp.place(x=512,y=40, anchor = CENTER), temp.after(2500, lambda: temp.place_forget())])
                save['font'] = font2
                save.place(x=512, y=550, anchor = CENTER)
            elif selectOrLogged == loggedMenu:
                temp=Label(self , text="Your changes have been saved!", background='white')
                temp['font'] = font2

                save2 = Button(self, text = "Save Changes", height=1, width=12, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [editRecipe(), temp.place(x=512,y=40, anchor = CENTER), temp.after(2500, lambda: temp.place_forget())])
                save2['font'] = font2
                save2.place(x=540, y=520)

                #add motor spin to this button
                dispense = Button(self, text = "Dispense", height=1, width=12, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [controller.show_frame(placeContainer)])
                # import main_func
                # command=lambda: [main_func.dispense(recipe)]
                dispense['font'] = font2
                dispense.place(x=300, y=520)
                table.e = Entry(table, width = 15, font=('Arial', 35))
                table.e.grid(row = r, column = 0)
                table.e.grid(row = r, column = 1)
                addSpice = Button(self, text = "+", height=1, width=3, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [morl(editMenu), controller.show_frame(selectMenu)])
                addSpice['font'] = font1
                addSpice.place(x=x, y=y, anchor = CENTER)

class positionMenu(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        font1 = tkFont.Font(family='Helvetica', size=15, weight ='bold')
        font2 = tkFont.Font(family='Helvetica', size=40, weight='bold')
        font3 = tkFont.Font(family='Helvetica', size=25, weight='bold')
        font4 = tkFont.Font(family='Helvetica', size=22, weight='bold')
        positionMenu.configure(self, background='white')

        goBack = Button(self, text = "⬅", height=1, width=7, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: controller.show_frame(mainMenu))
        goBack['font'] = font1
        goBack.place(x=20, y=10)

        Cont1 = Label(self, text = "Container 1: ", background='white')
        Cont1['font'] = font2
        Cont1.place(x=350, y=80, anchor=CENTER)

        Cont2 = Label(self, text = "Container 2: ", background='white')
        Cont2['font'] = font2
        Cont2.place(x=350, y=180, anchor=CENTER)

        Cont3 = Label(self, text = "Container 3: ", background='white')
        Cont3['font'] = font2
        Cont3.place(x=350, y=280, anchor=CENTER)

        Cont4 = Label(self, text = "Container 4: ", background='white')
        Cont4['font'] = font2
        Cont4.place(x=350, y=380, anchor=CENTER)

        Cont5 = Label(self, text = "Container 5: ", background='white')
        Cont5['font'] = font2
        Cont5.place(x=350, y=480, anchor=CENTER)

        self.option_add("*TCombobox*Listbox*Font", font3)

        ttk.Style().theme_use('default')
        style = ttk.Style() #If you dont have a class, put your root in the()
        style.configure('TCombobox', arrowsize=40, fieldbackground= "white", background = "white")
        style.configure('Vertical.TScrollbar', arrowsize=40)
        style.configure('TCombobox', postoffset=(0, 0, 0, -100))
        
        drop1 = ttk.Combobox(self, state = "readonly", width = 14)
        drop1['values']=positions
        drop1.current(0)
        drop1['font'] = font3
        drop1.place(x=675, y=80, anchor= CENTER)

        drop2 = ttk.Combobox(self, state = "readonly", width = 14)
        drop2['values']=positions
        drop2.current(1)
        drop2['font'] = font3
        drop2.place(x=675, y=180, anchor= CENTER)

        drop3 = ttk.Combobox(self, state = "readonly", width = 14)
        drop3['values']=positions
        drop3.current(2)
        drop3['font'] = font3
        drop3.place(x=675, y=280, anchor= CENTER)

        drop4 = ttk.Combobox(self, state = "readonly", width = 14)
        drop4['values']=positions
        drop4.current(3)
        drop4['font'] = font3
        drop4.place(x=675, y=380, anchor= CENTER)

        drop5 = ttk.Combobox(self, state = "readonly", width = 14)
        drop5['values']=positions
        drop5.current(4)
        drop5['font'] = font3
        drop5.place(x=675, y=480, anchor= CENTER)

        temp = Label(self , text = "", background='white')
        temp['font'] = font3

        save = Button(self, text = "Save Positions", height=1, width=12, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: [changePositions(drop1,drop2,drop3,drop4,drop5,temp), temp.place(x=512,y=25, anchor = CENTER), temp.after(2500, lambda: temp.place_forget())])
        save['font'] = font4
        save.place(x=512, y=550, anchor = CENTER)

class loggedMenu(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        loggedMenu.configure(self, background='white')

        font1= tkFont.Font(family='Helvetica', size=10, weight ='bold')
        font2= tkFont.Font(family='Helvetica', size=15, weight ='bold')
        
        goBack = Button(self, text = "⬅", height=1, width=7, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda: controller.show_frame(mainMenu))
        goBack['font'] = font2
        goBack.place(x=20, y=10)

        global savedRecipes
        savedRecipes = {}
        file = "Recipes.xlsx"
        workbook = openpyxl.load_workbook(file)
        sheetNames = workbook.sheetnames
        sheetNames.pop(0)
        spice = ""
        num = 0
        mes = ""
        for name in sheetNames:
            active = workbook[name]
            max_row = active.max_row
            max_col = active.max_column
            savedRecipes[name] = ""
            newRecipe = {}
            for i in range(1, max_row + 1):
                for j in range(1, max_col + 1):
                    cell = active.cell(row = i, column = j)
                    if j == 1:
                        spice = cell.value
                        newRecipe[spice] = ""
                    elif j == 2:
                        num = cell.value
                    elif j == 3:
                        mes = cell.value
                        quant = fracmes(num, mes)
                        newRecipe[spice] = quant
                savedRecipes[name] = newRecipe

        table = Label(self, relief="solid", background='white')
        table.place(x=75, y=75)

        
        #Create table
        r = 1
        c = 0

        #Column Headers
        table.e = Entry(table, width = 13, font=('Arial', 32, 'bold'))
        table.e.grid(row = 0,column = 0)
        table.e.insert(END, "RECIPE")
        table.e.configure(state = DISABLED, disabledbackground = 'white', disabledforeground= 'black')

        table.e = Entry(table, width = 20, font=('Arial', 32, 'bold'))
        table.e.grid(row = 0,column = 1)
        table.e.insert(END, "CONTENTS")
        table.e.configure(state = DISABLED, disabledbackground = 'white', disabledforeground= 'black')

        #for button
        x = 925
        y = 160
        
        #Format: RECIPE NAME | CONTENTS
        #        BBQ         | Salt(1/4 tsp.)
        #                    | Pepper(3/4 tbsp.)
        # {Recipe Name : {SpiceName : fracmes, SpiceName : fracmes}, Recipe Name: {SpiceName : fracmes, SpiceName : fracmes}}
        buttons = []
        index = 1
        for rName, spices in savedRecipes.items():
            table.e = Entry(table, width = 13, font=('Arial', 32))
            table.e.grid(row = r,column = c)
            table.e.insert(END, rName)
            table.e.configure(state = DISABLED, disabledbackground = 'white', disabledforeground= 'black')
            
            #Create button to edit here
            Select = Button(self, text = "Select", height=2, width=6, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda rName=rName: [setRecipe(rName), selected(rName), sorl(loggedMenu), morl(editMenu), controller.show_frame(editMenu)])
            Select.place(x=915, y=y, anchor = CENTER)
            Select['font'] = font1

            Delete = Button(self, text = "     🗑️", height=1, width=3, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda index=index: [deleteRecipe(index), controller.show_frame(loggedMenu)])
            Delete.place(x=975, y=y, anchor = CENTER)
            Delete['font'] = font2

            buttons.append(Delete)
            for key, val in spices.items():
                c += 1
                vfrac = fractions.Fraction(val.dec)
                table.e = Entry(table, width = 20, font=('Arial', 32))
                table.e.grid(row = r, column = c)
                table.e.insert(END, key + "(" + reduceFrac(vfrac) + " " + val.amt + ".)")
                table.e.configure(state = DISABLED, disabledbackground = 'white', disabledforeground= 'black')
                c = 0
                r += 1
                y+= 53
            index += 1

class placeContainer(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)

        placeContainer.configure(self, background='white')

        font1 = tkFont.Font(family='Helvetica', size=40, weight='bold')
        font2= tkFont.Font(family='Helvetica', size=22, weight ='bold')

        place = Label(self, text = "Please place container on scale \nand press continue.", background='white')
        place['font'] = font1
        place.place(x=512, y=200, anchor=CENTER)

        Continue = Button(self, text = "Continue", height=1, width=10, background='red', foreground='white', activebackground='maroon', activeforeground= 'white', command=lambda : [runDispense(place, Continue, controller, self)])
        Continue.place(x=512, y=550, anchor = CENTER)
        Continue['font'] = font2  

            
app = main()
app.mainloop()
